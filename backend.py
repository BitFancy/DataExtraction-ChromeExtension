from flask import Flask, jsonify, request
from flask_cors import CORS
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import Workbook
import json
import os

load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY)

wb = Workbook()
sheet = wb.active
title_items = ['Name', 'Company name']
for i in range(len(title_items)):
    sheet.cell(row=1, column=i + 1).value = title_items[i]

extracted_data = []

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": ["https://rephonic.com"], "methods": ["GET", "POST"]}})

@app.route('/get-fulltext', methods=['POST'])
def run_script():
    data = request.get_json()
    if data:
        global extracted_data
        print(f'Received data : {len(data['text'])}')

        prompt = data['text']
        gpt_response = function_calling(prompt)
        print(gpt_response)

        extracted_data.append(json.loads(gpt_response))

        with open('gpt_response.txt', 'a', encoding='utf-8') as text:
            text.write('Prompt ---> ' + prompt + '\n' + 'Extrated data ---> ' + gpt_response + '\n\n')

        start_row = 2
        for item in extracted_data:
            print(item)
            sheet.cell(row=start_row, column=1).value = item['name']
            sheet.cell(row=start_row, column=2).value = item['company']
            wb.save('final_result.xlsx')
            start_row += 1

        response = {
            "status": "success",
            "message": "Data received successfully!",
            "result": json.loads(gpt_response)
        }

        return jsonify(response), 200
    else:
        response = {
            "status": "failure",
            "message": "No data received"
        }
        return jsonify(response), 400
    
def function_calling(prompt):
    extract_info = [
        {
            "type": "function",
            "function": {
                "name": "extracte_info",
                "description": "Extracts the full name of the guest and the company name from given text.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "The full name of the guest"
                        },
                        "company": {
                            "type": "string",
                            "description": "The name of the guest's company"
                        },
                    },
                    "required": ["name", "company"]
                }
            }
        }
    ]

    messages = [
        {"role": "system", "content": "You are an expert at extracting specific information from text. Please extract the essential information from the data that is inputted by user. You should answer in all of question."},
        {"role": "user", "content": prompt}
        ]

    response = client.chat.completions.create(
        model="gpt-4-1106-preview",
        messages=messages,
        temperature=0,
        tools=extract_info,
    )

    return response.choices[0].message.tool_calls[0].function.arguments


if __name__ == "__main__":
    app.run(port=5002, debug=True)